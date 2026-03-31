#!/usr/bin/env python3
"""
Process voksmonitor CSV export and generate three output tables:
1. answers.csv - rows are sessions, columns are questions, values are answers
2. candidates_match.csv - rows are sessions, columns are candidates, values are match percentages
3. organizations_match.csv - rows are sessions, columns are organizations, values are match percentages

Also generates an xlsx file with all three tables as separate sheets.
"""

import csv
import json
from pathlib import Path
from typing import Dict, List, Any
from collections import defaultdict
from openpyxl import Workbook


def load_json_file(filepath: Path) -> List[Dict[str, Any]]:
    """Load and parse a JSON file."""
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def get_candidate_to_organization_mapping(
    candidates: List[Dict], organizations: List[Dict]
) -> Dict[str, str]:
    """Create a mapping from candidate ID to organization ID."""
    mapping = {}

    def process_candidate(candidate: Dict):
        """Recursively process candidates and nested candidates."""
        candidate_id = candidate["id"]
        # Get organization from references
        for ref in candidate.get("references", []):
            if ref["type"] == "organization":
                mapping[candidate_id] = ref["id"]
                break

        # Process nested candidates
        for nested in candidate.get("nestedCandidates", []):
            nested_id = nested["id"]
            # Nested candidates inherit organization from parent
            if candidate_id in mapping:
                mapping[nested_id] = mapping[candidate_id]

    for candidate in candidates:
        process_candidate(candidate)

    return mapping


def get_organization_name(org_id: str, organizations: List[Dict]) -> str:
    """Get organization short name by ID."""
    for org in organizations:
        if org["id"] == org_id:
            return org.get("shortName", org["name"])
    return org_id


def get_question_title(question_id: str, questions: List[Dict]) -> str:
    """Get question title by ID."""
    for q in questions:
        if q["id"] == question_id:
            return q.get("statement", question_id)
    return question_id


def get_candidate_name(
    candidate_id: str,
    candidates: List[Dict],
    persons: List[Dict],
    organizations: List[Dict],
) -> str:
    """Get candidate name by ID.

    For top-level candidates (organizations), returns organization short name.
    For nested candidates (persons), returns person name.
    """
    # Build a lookup for all candidates (both top-level and nested)
    for candidate in candidates:
        if candidate["id"] == candidate_id:
            # Top-level candidate - use organization name or displayName
            if "displayName" in candidate:
                return candidate["displayName"]
            for ref in candidate.get("references", []):
                if ref["type"] == "organization":
                    return get_organization_name(ref["id"], organizations)
            return candidate_id

        # Check nested candidates
        for nested in candidate.get("nestedCandidates", []):
            if nested["id"] == candidate_id:
                # Nested candidate - use person name
                for ref in nested.get("references", []):
                    if ref["type"] == "person":
                        for person in persons:
                            if person["id"] == ref["id"]:
                                return person["name"]
                return candidate_id

    return candidate_id


def process_csv_export(csv_path: Path, output_dir: Path):
    """Process the CSV export and generate output tables."""

    # Load reference data
    base_dir = "2026-ogy/inventory"
    questions = load_json_file(Path(base_dir) / "questions.json")
    candidates = load_json_file(Path(base_dir) / "candidates.json")
    organizations = load_json_file(Path(base_dir) / "organizations.json")
    persons = load_json_file(Path(base_dir) / "persons.json")

    # Create mappings
    candidate_to_org = get_candidate_to_organization_mapping(candidates, organizations)

    # Load demography data and build a lookup by sessionId
    demog_path = Path(__file__).parent / "voksmonitor_demo_202603301606.csv"
    demog_lookup = {}
    demog_columns = []
    if demog_path.exists():
        with open(demog_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            demog_columns = [
                col
                for col in reader.fieldnames
                if col not in ("id", "sessionId", "calculatorId", "calculatorKey")
            ]
            for row in reader:
                sid = row.get("sessionId")
                if sid:
                    demog_lookup[sid] = row
    else:
        print(f"WARNING: Demography file not found at {demog_path}")

    # Data structures to collect data
    sessions_data = []
    all_question_ids = set()
    all_candidate_ids = set()
    all_org_ids = set()
    # No need to track questions_with_isimportant for this approach

    from datetime import datetime, timezone, timedelta

    # Parse CSV
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            session_id = row["sessionId"]
            if session_id in [
                "76c09c82-1d72-4851-9165-f651cb00d26f",
                "48e17148-3e14-4afa-86da-78f398981f82",
                "962e3067-938b-44fc-b58a-33c844523755",
                "f77807f0-799c-42c8-9df9-ae8507683814",
                "71127f8c-728d-4498-8187-df875cf1e340",
            ]:
                continue  # Skip test sessions

            # Filter by createdAt
            created_at_str = row.get("createdAt", "")
            if not created_at_str:
                continue
            try:
                # Example: '2026-03-27 12:12:37.067 +0100'
                created_at = datetime.strptime(
                    created_at_str, "%Y-%m-%d %H:%M:%S.%f %z"
                )
            except Exception:
                continue
            # Only include if after 2026-03-27 08:30 (local time +0100)
            filter_dt = datetime(2026, 3, 27, 8, 30, tzinfo=created_at.tzinfo)
            if created_at <= filter_dt:
                continue

            # Parse answers
            answers_json = row.get("answers", "[]")
            if not answers_json:
                answers_json = "[]"

            try:
                answers = json.loads(answers_json)
            except json.JSONDecodeError:
                answers = []

            # Parse results (matches)
            result_json = row.get("result", "[]")
            if not result_json:
                result_json = "[]"

            try:
                results = json.loads(result_json)
            except json.JSONDecodeError:
                results = []

            # Process answers
            answers_dict = {}
            isimportant_dict = {}
            for answer in answers:
                question_id = answer.get("questionId")
                if question_id:
                    all_question_ids.add(question_id)
                    # Store answer value: True, False, or None (if not answered)
                    answer_value = answer.get("answer")
                    answers_dict[question_id] = answer_value
                    # Store isImportant if present
                    if "isImportant" in answer:
                        isimportant_dict[question_id] = answer["isImportant"]

            # Process candidate matches
            candidate_matches = {}
            org_matches = defaultdict(list)

            for result in results:
                candidate_id = result.get("id")
                match_value = result.get("match")

                if candidate_id and match_value is not None:
                    all_candidate_ids.add(candidate_id)
                    candidate_matches[candidate_id] = match_value

                    # Map to organization
                    org_id = candidate_to_org.get(candidate_id)
                    if org_id:
                        all_org_ids.add(org_id)
                        org_matches[org_id].append(match_value)

            # Calculate average match per organization
            org_avg_matches = {}
            for org_id, matches in org_matches.items():
                org_avg_matches[org_id] = (
                    sum(matches) / len(matches) if matches else None
                )

            sessions_data.append(
                {
                    "session_id": session_id,
                    "createdAt": row.get("createdAt", ""),
                    "updatedAt": row.get("updatedAt", ""),
                    "completedAt": row.get("completedAt", ""),
                    "answers": answers_dict,
                    "isimportant": isimportant_dict,
                    "candidate_matches": candidate_matches,
                    "org_matches": org_avg_matches,
                }
            )

    # Sort IDs for consistent column ordering
    question_ids_sorted = sorted(all_question_ids)
    candidate_ids_sorted = sorted(all_candidate_ids)
    org_ids_sorted = sorted(all_org_ids)

    # Filter out questions without a title and log warnings
    valid_question_ids = []
    for qid in question_ids_sorted:
        title = get_question_title(qid, questions)
        if title == qid:
            print(f"WARNING: Skipping question with id '{qid}' (no title found)")
        else:
            valid_question_ids.append(qid)

    # Generate answers.csv
    output_dir.mkdir(parents=True, exist_ok=True)
    with open(output_dir / "answers.csv", "w", encoding="utf-8", newline="") as f:
        # Create header with question titles (only valid questions)
        headers = (
            ["session_id", "createdAt", "updatedAt", "completedAt"]
            + demog_columns
            + [get_question_title(qid, questions) for qid in valid_question_ids]
        )
        writer = csv.writer(f)
        writer.writerow(headers)

        for session in sessions_data:
            row = [
                session["session_id"],
                session["createdAt"],
                session["updatedAt"],
                session["completedAt"],
            ]
            demog = demog_lookup.get(session["session_id"], {})
            for col in demog_columns:
                row.append(demog.get(col, ""))
            for qid in valid_question_ids:
                answer = session["answers"].get(qid)
                isimp = session.get("isimportant", {}).get(qid)
                if answer is True:
                    val = "True"
                elif answer is False:
                    val = "False"
                else:
                    val = ""
                if val and isimp is True:
                    val += " Important"
                row.append(val)
            writer.writerow(row)

    # Generate candidates_match.csv
    with open(
        output_dir / "candidates_match.csv", "w", encoding="utf-8", newline=""
    ) as f:
        headers = (
            ["session_id", "createdAt", "updatedAt", "completedAt"]
            + demog_columns
            + [
                get_candidate_name(cid, candidates, persons, organizations)
                for cid in candidate_ids_sorted
            ]
        )
        writer = csv.writer(f)
        writer.writerow(headers)

        for session in sessions_data:
            row = [
                session["session_id"],
                session["createdAt"],
                session["updatedAt"],
                session["completedAt"],
            ]
            demog = demog_lookup.get(session["session_id"], {})
            for col in demog_columns:
                row.append(demog.get(col, ""))
            for cid in candidate_ids_sorted:
                match_value = session["candidate_matches"].get(cid)
                row.append(f"{match_value:.2f}" if match_value is not None else "")
            writer.writerow(row)

    # Generate organizations_match.csv
    with open(
        output_dir / "organizations_match.csv", "w", encoding="utf-8", newline=""
    ) as f:
        headers = (
            ["session_id", "createdAt", "updatedAt", "completedAt"]
            + demog_columns
            + [get_organization_name(oid, organizations) for oid in org_ids_sorted]
        )
        writer = csv.writer(f)
        writer.writerow(headers)

        for session in sessions_data:
            row = [
                session["session_id"],
                session["createdAt"],
                session["updatedAt"],
                session["completedAt"],
            ]
            demog = demog_lookup.get(session["session_id"], {})
            for col in demog_columns:
                row.append(demog.get(col, ""))
            for oid in org_ids_sorted:
                match_value = session["org_matches"].get(oid)
                row.append(f"{match_value:.2f}" if match_value is not None else "")
            writer.writerow(row)

    print(
        f"✓ Generated answers.csv with {len(sessions_data)} rows and {len(valid_question_ids)} question columns"
    )
    print(
        f"✓ Generated candidates_match.csv with {len(sessions_data)} rows and {len(candidate_ids_sorted)} candidate columns"
    )
    print(
        f"✓ Generated organizations_match.csv with {len(sessions_data)} rows and {len(org_ids_sorted)} organization columns"
    )

    # Generate xlsx file with all three sheets
    wb = Workbook()

    # Answers sheet
    ws_answers = wb.active
    ws_answers.title = "Answers"
    answers_headers = (
        ["session_id", "createdAt", "updatedAt", "completedAt"]
        + demog_columns
        + [get_question_title(qid, questions) for qid in valid_question_ids]
    )
    ws_answers.append(answers_headers)
    for session in sessions_data:
        row = [
            session["session_id"],
            session["createdAt"],
            session["updatedAt"],
            session["completedAt"],
        ]
        demog = demog_lookup.get(session["session_id"], {})
        for col in demog_columns:
            row.append(demog.get(col, ""))
        for qid in valid_question_ids:
            answer = session["answers"].get(qid)
            isimp = session.get("isimportant", {}).get(qid)
            if answer is True:
                val = "True"
            elif answer is False:
                val = "False"
            else:
                val = ""
            if val and isimp is True:
                val += " Important"
            row.append(val)
        ws_answers.append(row)

    # Candidates Match sheet
    # ws_candidates = wb.create_sheet("Candidates Match")
    # candidates_headers = (
    #     ["session_id", "createdAt", "updatedAt", "completedAt"]
    #     + demog_columns
    #     + [
    #         get_candidate_name(cid, candidates, persons, organizations)
    #         for cid in candidate_ids_sorted
    #     ]
    # )
    # ws_candidates.append(candidates_headers)
    # for session in sessions_data:
    #     row = [
    #         session["session_id"],
    #         session["createdAt"],
    #         session["updatedAt"],
    #         session["completedAt"],
    #     ]
    #     demog = demog_lookup.get(session["session_id"], {})
    #     for col in demog_columns:
    #         row.append(demog.get(col, ""))
    #     for cid in candidate_ids_sorted:
    #         match_value = session["candidate_matches"].get(cid)
    #         row.append(round(match_value, 2) if match_value is not None else "")
    #     ws_candidates.append(row)

    # Organizations Match sheet
    ws_orgs = wb.create_sheet("Organizations Match")
    orgs_headers = (
        ["session_id", "createdAt", "updatedAt", "completedAt"]
        + demog_columns
        + [get_organization_name(oid, organizations) for oid in org_ids_sorted]
    )
    ws_orgs.append(orgs_headers)
    for session in sessions_data:
        row = [
            session["session_id"],
            session["createdAt"],
            session["updatedAt"],
            session["completedAt"],
        ]
        demog = demog_lookup.get(session["session_id"], {})
        for col in demog_columns:
            row.append(demog.get(col, ""))
        for oid in org_ids_sorted:
            match_value = session["org_matches"].get(oid)
            row.append(round(match_value, 2) if match_value is not None else "")
        ws_orgs.append(row)

    xlsx_path = output_dir / "voksmonitor_export.xlsx"
    wb.save(xlsx_path)
    print(f"✓ Generated voksmonitor_export.xlsx with 3 sheets")


if __name__ == "__main__":
    csv_path = Path(__file__).parent / "voksmonitor_data_202603301601.csv"
    output_dir = Path(__file__).parent

    if not csv_path.exists():
        print(f"Error: CSV file not found at {csv_path}")
        exit(1)

    process_csv_export(csv_path, output_dir)
    print("\nDone! Output files created in the same directory.")
